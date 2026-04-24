from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse, reverse_lazy
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q
from django.views.generic import TemplateView, ListView, UpdateView, DetailView, CreateView, DeleteView, View
from django.http import HttpResponse
from django.db import transaction
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from django.contrib.auth.models import User


from ..models import Activo, Edificio, Piso, Ubicacion, Marca, Categoria, Estado, Catalogo, Funcionario, AuditoriaActivo
from ..forms import ActivoForm
from ..utils import _get_excel_val
from ..mixins import GroupRequiredMixin


# ---------------------------------------    
# NUEVAS VISTAS
# ---------------------------------------
class InicioNuevoView(LoginRequiredMixin, GroupRequiredMixin, TemplateView):
    """
    Vista de inicio (Requiere Login)
    """
    template_name = 'home_nuevo.html'
    group_required = ['ADR', 'Alumno en Práctica', 'Auxiliar Operador ADR', 'Operador ADR']




class ListaActivosView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    """
    Vista unificada para listar activos con filtros avanzados por Query Params.
    Reemplaza a activo_list.
    """
    model = Activo
    template_name = 'lista_activos.html'
    paginate_by = 20
    context_object_name = 'activos'  # ListView utiliza 'page_obj' por defecto cuando hay paginación
    group_required = ['ADR', 'Alumno en Práctica', 'Auxiliar Operador ADR', 'Operador ADR']
    
    def get_queryset(self):
        # 1. Optimización de la consulta con select_related
        queryset = Activo.objects.select_related(
            'catalogo__categoria', 'catalogo__marca', 
            'ubicacion__piso__edificio', 'asignado_a', 'estado', 'creado_por'
        ).all().order_by('-updated_at')
        
        # 2. Captura de parámetros GET y guardado en la instancia para usar en contexto
        self.search_query = self.request.GET.get('search', '')
        self.categoria_id = self.request.GET.get('categoria', '')
        self.categoria_nombre = self.request.GET.get('categoria_nombre', '')
        self.marca_id = self.request.GET.get('marca', '')
        self.estado_id = self.request.GET.get('estado', '')
        self.edificio_id = self.request.GET.get('edificio', '')
        self.ubicacion_nombre = self.request.GET.get('ubicacion_nombre', '')
        self.tipo_uso = self.request.GET.get('tipo_uso', '')
        self.tipo_red = self.request.GET.get('tipo_red', '')

        self.titulo_lista = "Listado General de Activos"

        # 3. Aplicación dinámica de filtros
        if self.categoria_id:
            queryset = queryset.filter(catalogo__categoria_id=self.categoria_id)
            try:
                cat_obj = Categoria.objects.get(id=self.categoria_id)
                self.titulo_lista = f"Inventario de {cat_obj.nombre}s"
            except Categoria.DoesNotExist:
                pass

        if self.categoria_nombre:
            queryset = queryset.filter(catalogo__categoria__nombre__icontains=self.categoria_nombre)
            self.titulo_lista = f"Inventario de {self.categoria_nombre}s"

        if self.marca_id:
            queryset = queryset.filter(catalogo__marca_id=self.marca_id)
            
        if self.estado_id:
            queryset = queryset.filter(estado_id=self.estado_id)
            
        if self.edificio_id:
            queryset = queryset.filter(ubicacion__piso__edificio_id=self.edificio_id)

        if self.ubicacion_nombre:
            queryset = queryset.filter(ubicacion__nombre__icontains=self.ubicacion_nombre)
            self.titulo_lista = f"Equipos en {self.ubicacion_nombre}"
            
        if self.tipo_uso:
            queryset = queryset.filter(tipo_uso=self.tipo_uso)
            if self.tipo_uso == Activo.TipoUso.EVENTOS:
                self.titulo_lista = "Equipos para Eventos"

        if self.tipo_red:
            queryset = queryset.filter(tipo_red=self.tipo_red)
            if self.tipo_red == Activo.TipoRed.ISLA:
                self.titulo_lista = "Equipos Isla"

        # 4. Búsqueda de texto global
        if self.search_query:
            queryset = queryset.filter(
                Q(numero_serie__icontains=self.search_query) |
                Q(etiqueta__icontains=self.search_query) |
                Q(bdo__icontains=self.search_query) |
                Q(netbios__icontains=self.search_query) |
                Q(asignado_a__nombre__icontains=self.search_query) |
                Q(catalogo__modelo__icontains=self.search_query)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Preservar query params para la paginación
        get_copy = self.request.GET.copy()
        if 'page' in get_copy:
            get_copy.pop('page')
        query_string = get_copy.urlencode()

        context.update({
            'titulo_lista': getattr(self, 'titulo_lista', "Listado General de Activos"),
            'query_string': f"&{query_string}" if query_string else "",
            
            # Listas para rellenar los select del HTML
            'categorias': Categoria.objects.all().order_by('nombre'),
            'marcas': Marca.objects.all().order_by('nombre'),
            'estados': Estado.objects.all().order_by('nombre'),
            'edificios': Edificio.objects.all().order_by('nombre'),
            'tipos_uso': Activo.TipoUso.choices,
            
            # Variables para mantener seleccionada la opción correcta en la vista
            'search_query': getattr(self, 'search_query', ''),
            'categoria_seleccionada': getattr(self, 'categoria_id', ''),
            'marca_seleccionada': getattr(self, 'marca_id', ''),
            'estado_seleccionado': getattr(self, 'estado_id', ''),
            'edificio_seleccionado': getattr(self, 'edificio_id', ''),
            'tipo_uso_seleccionado': getattr(self, 'tipo_uso', ''),
        })
        return context




class EditarActivoView(LoginRequiredMixin, GroupRequiredMixin, UpdateView):
    """
    Vista unificada para editar cualquier tipo de activo.
    """
    model = Activo
    form_class = ActivoForm
    template_name = 'editar_activo.html'
    context_object_name = 'activo'
    group_required = ['ADR', 'Auxiliar Operador ADR', 'Operador ADR']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # self.object es la instancia del activo que estamos editando
        context['titulo_form'] = f"Editar {self.object.catalogo.categoria.nombre}"
        return context

    def form_valid(self, form):
        # Dispara el guardado y emite el mensaje de éxito
        response = super().form_valid(form)
        messages.success(self.request, f'El activo {self.object} ha sido actualizado correctamente.')
        return response

    def form_invalid(self, form):
        # Emite el mensaje de error si la validación falla
        messages.error(self.request, 'No se pudo guardar. Por favor, corrige los errores en el formulario.')
        return super().form_invalid(form)

    def get_success_url(self):
        return reverse_lazy('lista_activos')




class DetalleActivoView(LoginRequiredMixin, GroupRequiredMixin, DetailView):
    """
    Vista para mostrar el detalle completo de un activo.
    """
    model = Activo
    template_name = 'ver_activo.html'
    context_object_name = 'activo'
    group_required = ['ADR', 'Alumno en Práctica', 'Auxiliar Operador ADR', 'Operador ADR']

    def get_queryset(self):
        # Pre-cargamos todas las relaciones para que la vista sea rápida y eficiente
        return Activo.objects.select_related(
            'catalogo__categoria', 'catalogo__marca', 
            'estado', 'ubicacion__piso__edificio', 
            'asignado_a__cargo', 'asignado_a__area'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_detalle'] = f"Detalle del Equipo: {self.object.catalogo.marca} {self.object.catalogo.modelo or ''}".strip()
        return context




class AgregarActivoView(LoginRequiredMixin, GroupRequiredMixin, CreateView):
    """
    Vista para registrar un nuevo activo.
    """
    model = Activo
    form_class = ActivoForm
    template_name = 'agregar_activo.html'
    group_required = ['ADR', 'Auxiliar Operador ADR', 'Operador ADR']

    def get_form_kwargs(self):
        # Pasamos el parámetro de categoría al formulario, al igual que en la vista original
        kwargs = super().get_form_kwargs()
        kwargs['categoria_nombre'] = self.request.GET.get('categoria_nombre', '')
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        categoria_nombre = self.request.GET.get('categoria_nombre', '')
        
        context['titulo_form'] = f"Nuevo Activo: {categoria_nombre}" if categoria_nombre else "Registrar Nuevo Activo"
        context['categoria_nombre'] = categoria_nombre
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'Activo registrado exitosamente: {self.object}')
        return response

    def form_invalid(self, form):
        messages.error(self.request, 'No se pudo guardar el activo. Por favor, corrige los errores del formulario.')
        return super().form_invalid(form)

    def get_success_url(self):
        # Redirigir de vuelta a la lista con el mismo filtro de la categoría
        redirect_url = reverse('lista_activos')
        categoria_nombre = self.request.GET.get('categoria_nombre', '')
        if categoria_nombre:
            redirect_url += f"?categoria_nombre={categoria_nombre}"
        return redirect_url


class EliminarActivoView(LoginRequiredMixin, GroupRequiredMixin, DeleteView):
    """
    Vista para procesar la eliminación (soft-delete) de un activo.
    """
    model = Activo
    template_name = 'eliminar_activo.html'
    context_object_name = 'activo'
    success_url = reverse_lazy('lista_activos')
    group_required = ['ADR', 'Operador ADR']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_form'] = 'Confirmar Eliminación'
        return context

    def form_valid(self, form):
        # En Django moderno (4.0+), form_valid maneja la lógica de DeleteView
        activo = self.get_object()
        # Llamar a delete() invoca override en adr/models.py (soft-delete)
        activo.delete()
        messages.success(self.request, f'El equipo {activo} ha sido enviado a la papelera (Eliminado).')
        return redirect(self.success_url)




class SubirExcelActivosView(LoginRequiredMixin, GroupRequiredMixin, View):
    """
    Vista para importar activos masivamente mediante Excel.
    Aplica reglas estrictas y mapea etiquetas legibles de vuelta a sus códigos internos.
    """
    template_name = 'subir_excel_activos.html'
    group_required = ['ADR', 'Auxiliar Operador ADR', 'Operador ADR']

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        if 'archivo_excel' not in request.FILES:
            messages.error(request, 'Por favor, selecciona un archivo Excel válido.')
            return redirect('upload_excel_activos')
            
        archivo = request.FILES['archivo_excel']
        
        if not archivo.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'El formato del archivo no es válido. Debe ser .xlsx o .xls')
            return redirect('upload_excel_activos')
            
        try:
            # sheet_name=0 asegura que siempre lea la primera hoja ("Equipos")
            df = pd.read_excel(archivo, sheet_name=0)
            
            # Limpiamos nombres de columnas (quitar espacios y poner a mayúsculas)
            df.columns = df.columns.str.strip().str.upper()
            
            # Reemplazar valores NaN por None
            df = df.where(pd.notnull(df), None)

            # MAPEO INVERSO: Diccionarios para traducir etiqueta legible (ej: "Red Institucional (Dominio)") -> Código (ej: "DOM")
            red_map = {str(label).strip().upper(): key for key, label in Activo.TipoRed.choices}
            uso_map = {str(label).strip().upper(): key for key, label in Activo.TipoUso.choices}
            
            registros_exitosos = 0
            errores = []

            with transaction.atomic():
                for index, row in df.iterrows():
                    try:
                        # 1. Validar Categoría
                        cat_nombre = _get_excel_val(row, 'CATEGORIA', default='', to_upper=True)
                        if not cat_nombre:
                            errores.append(f"Fila {index + 2}: La categoría es obligatoria.")
                            continue
                            
                        try:
                            categoria = Categoria.objects.get(nombre__iexact=cat_nombre)
                        except Categoria.DoesNotExist:
                            errores.append(f"Fila {index + 2}: La categoría '{cat_nombre}' no existe en el sistema. Registro denegado.")
                            continue

                        # 2. Validar Ubicación
                        ubicacion_nombre = _get_excel_val(row, 'UBICACION', default='')
                        ubicacion_obj = None
                        if ubicacion_nombre:
                            nombre_edi = _get_excel_val(row, 'EDIFICIO', default='')
                            nombre_piso = _get_excel_val(row, 'PISO', default='')

                            if not nombre_edi or not nombre_piso:
                                errores.append(f"Fila {index + 2}: Para asignar la ubicación '{ubicacion_nombre}', debe especificar EDIFICIO y PISO.")
                                continue

                            ubicacion_obj = Ubicacion.objects.filter(
                                nombre__iexact=ubicacion_nombre,
                                piso__nombre__iexact=nombre_piso,
                                piso__edificio__nombre__iexact=nombre_edi
                            ).first()
                            
                            if not ubicacion_obj:
                                errores.append(f"Fila {index + 2}: La ubicación '{ubicacion_nombre}' (Piso: {nombre_piso}, Edificio: {nombre_edi}) no existe en el sistema. Registro denegado.")
                                continue

                        # 3. Procesar Marca y Modelo
                        marca_nombre = _get_excel_val(row, 'MARCA', default='', to_upper=True)
                        if not marca_nombre:
                            errores.append(f"Fila {index + 2}: La marca es obligatoria.")
                            continue
                            
                        modelo_nombre = _get_excel_val(row, 'MODELO', default='GENÉRICO', to_upper=True)
                        
                        marca, _ = Marca.objects.get_or_create(nombre=marca_nombre)
                        catalogo, _ = Catalogo.objects.get_or_create(
                            categoria=categoria, 
                            marca=marca, 
                            modelo=modelo_nombre
                        )

                        # 4. Procesar Estado y Asignatario
                        estado_nombre = _get_excel_val(row, 'ESTADO', default='OPERATIVO', to_upper=True)
                        estado, _ = Estado.objects.get_or_create(nombre=estado_nombre)

                        funcionario_nombre = _get_excel_val(row, 'ASIGNATARIO', default='', to_upper=True)
                        funcionario_obj = None
                        if funcionario_nombre:
                            funcionario_obj, _ = Funcionario.objects.get_or_create(nombre=funcionario_nombre)

                        # 5. Mapeo de Red y Uso desde Etiquetas Legibles
                        tipo_red_label = _get_excel_val(row, 'TIPO_RED', default='', to_upper=True)
                        tipo_uso_label = _get_excel_val(row, 'TIPO_USO', default='', to_upper=True)

                        # Buscamos la etiqueta en nuestro mapa. Si no existe/es inválida, usa defaults ('DOM' y 'PER')
                        tipo_red = red_map.get(tipo_red_label, 'DOM')
                        tipo_uso = uso_map.get(tipo_uso_label, 'PER')

                        # 6. Creación/Actualización del Activo
                        numero_serie = _get_excel_val(row, 'NUMERO_SERIE', default=None)
                        etiqueta = _get_excel_val(row, 'ETIQUETA', default=None)
                        bdo = _get_excel_val(row, 'BDO', default=None)
                        netbios = _get_excel_val(row, 'NETBIOS', default=None)

                        activo = None
                        if bdo:
                            activo = Activo.objects.filter(bdo=bdo).first()
                        if not activo and numero_serie:
                            activo = Activo.objects.filter(numero_serie=numero_serie).first()

                        if activo:
                            activo.catalogo = catalogo
                            activo.estado = estado
                            activo.ubicacion = ubicacion_obj
                            activo.asignado_a = funcionario_obj
                            activo.tipo_red = tipo_red
                            activo.tipo_uso = tipo_uso
                            if netbios: activo.netbios = netbios
                            activo.save()
                        else:
                            nuevo_activo = Activo(
                                catalogo=catalogo,
                                estado=estado,
                                numero_serie=numero_serie,
                                etiqueta=etiqueta,
                                bdo=bdo,
                                netbios=netbios,
                                tipo_red=tipo_red,
                                tipo_uso=tipo_uso,
                                ubicacion=ubicacion_obj,
                                asignado_a=funcionario_obj
                            )
                            nuevo_activo.save()
                            
                        registros_exitosos += 1

                    except Exception as e:
                        errores.append(f"Fila {index + 2}: Error interno al procesar ({str(e)})")

            if errores:
                for error in errores[:10]:
                    messages.warning(request, error)
                if len(errores) > 10:
                    messages.warning(request, f"...y {len(errores) - 10} errores más omitidos.")

            if registros_exitosos > 0:
                messages.success(request, f'Proceso completado. Se importaron/actualizaron {registros_exitosos} activos.')
            else:
                messages.info(request, 'El proceso finalizó, pero no se importó ningún registro nuevo.')
                
            return redirect('lista_activos')

        except Exception as e:
            messages.error(request, f'Error al leer el archivo Excel: {str(e)}')
            return redirect('subir_excel_activos')




class DescargarPlantillaExcelView(LoginRequiredMixin, GroupRequiredMixin, View):
    """
    Genera un archivo Excel (.xlsx) con las cabeceras correctas y listas desplegables
    basadas en los datos actuales del sistema (incluyendo etiquetas legibles para choices).
    """
    group_required = ['ADR', 'Alumno en Práctica', 'Auxiliar Operador ADR', 'Operador ADR']

    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        
        # --- HOJA 1: EQUIPOS ---
        ws_equipos = wb.active
        ws_equipos.title = "Equipos"
        
        cabeceras = [
            'CATEGORIA', 'MARCA', 'MODELO', 'NUMERO_SERIE', 'ETIQUETA', 
            'BDO', 'NETBIOS', 'TIPO_RED', 'TIPO_USO', 'ESTADO', 
            'EDIFICIO', 'PISO', 'UBICACION', 'ASIGNATARIO'
        ]
        
        header_fill = PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, header_title in enumerate(cabeceras, 1):
            cell = ws_equipos.cell(row=1, column=col_num, value=header_title)
            cell.fill = header_fill
            cell.font = header_font
            ws_equipos.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

        # --- HOJA 2: DATOS ---
        ws_datos = wb.create_sheet(title="Datos")
        
        # 1. Obtener datos de la BD
        categorias = list(Categoria.objects.values_list('nombre', flat=True).order_by('nombre'))
        estados = list(Estado.objects.values_list('nombre', flat=True).order_by('nombre'))
        edificios = list(Edificio.objects.values_list('nombre', flat=True).order_by('nombre'))
        pisos = list(Piso.objects.values_list('nombre', flat=True).order_by('nombre'))
        ubicaciones = list(Ubicacion.objects.values_list('nombre', flat=True).order_by('nombre'))
        marcas = list(Marca.objects.values_list('nombre', flat=True).order_by('nombre'))
        
        # IMPORTANTE: Ahora extraemos index 1 (Label humano) en lugar de index 0 (Código de BD)
        tipos_red_labels = [choice[1] for choice in Activo.TipoRed.choices]
        tipos_uso_labels = [choice[1] for choice in Activo.TipoUso.choices]

        # 2. Escribir datos
        datos_diccionario = {
            1: categorias,        # Columna A
            2: estados,           # Columna B
            3: tipos_red_labels,  # Columna C (Humanos)
            4: tipos_uso_labels,  # Columna D (Humanos)
            5: edificios,         # Columna E
            6: pisos,             # Columna F
            7: ubicaciones,       # Columna G
            8: marcas             # Columna H (NUEVO: Marcas)
        }

        for col_index, valores in datos_diccionario.items():
            for row_index, valor in enumerate(valores, 1):
                ws_datos.cell(row=row_index, column=col_index, value=valor)

        ws_datos.sheet_state = 'hidden'

        # --- APLICAR VALIDACIÓN DE DATOS (Desplegables) ---
        def crear_dropdown(col_letra_datos, total_filas, col_letra_equipos, allow_blank=True):
            if total_filas > 0:
                formula = f"Datos!${col_letra_datos}$1:${col_letra_datos}${total_filas}"
                # Para marcas, permitimos espacios en blanco/escribir advertencias
                dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank, showErrorMessage=False)
                dv.add(f"{col_letra_equipos}2:{col_letra_equipos}1000")
                ws_equipos.add_data_validation(dv)

        crear_dropdown('A', len(categorias), 'A')        # A = CATEGORIA
        crear_dropdown('B', len(estados), 'J')            # J = ESTADO
        crear_dropdown('C', len(tipos_red_labels), 'H')   # H = TIPO_RED
        crear_dropdown('D', len(tipos_uso_labels), 'I')   # I = TIPO_USO
        crear_dropdown('E', len(edificios), 'K')          # K = EDIFICIO
        crear_dropdown('F', len(pisos), 'L')              # L = PISO
        crear_dropdown('G', len(ubicaciones), 'M')        # M = UBICACION
        crear_dropdown('H', len(marcas), 'B')             # B = MARCA (NUEVO)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Plantilla_Ingreso_Activos.xlsx"'
        
        wb.save(response)
        return response




class DescargarExcelActivosView(LoginRequiredMixin, GroupRequiredMixin, View):
    """
    Vista para exportar los activos registrados en el sistema a un archivo Excel.
    Aplica de manera dinámica los mismos filtros que la vista de listado.
    """
    group_required = ['ADR', 'Alumno en Práctica', 'Auxiliar Operador ADR', 'Operador ADR']

    def get(self, request, *args, **kwargs):
        # 1. Obtenemos todos los activos optimizando las relaciones
        activos = Activo.objects.select_related(
            'catalogo__categoria', 'catalogo__marca', 'estado',
            'ubicacion__piso__edificio', 'asignado_a__cargo'
        ).all().order_by('-updated_at')

        # Capturamos los mismos Query Params que usa la lista de activos
        search_query = request.GET.get('search', '')
        categoria_id = request.GET.get('categoria', '')
        categoria_nombre = request.GET.get('categoria_nombre', '')
        marca_id = request.GET.get('marca', '')
        estado_id = request.GET.get('estado', '')
        edificio_id = request.GET.get('edificio', '')
        ubicacion_nombre = request.GET.get('ubicacion_nombre', '')
        tipo_uso = request.GET.get('tipo_uso', '')
        tipo_red = request.GET.get('tipo_red', '')

        # Aplicamos los filtros dinámicamente si existen en la URL
        if categoria_id:
            activos = activos.filter(catalogo__categoria_id=categoria_id)
        if categoria_nombre:
            activos = activos.filter(catalogo__categoria__nombre__icontains=categoria_nombre)
        if marca_id:
            activos = activos.filter(catalogo__marca_id=marca_id)
        if estado_id:
            activos = activos.filter(estado_id=estado_id)
        if edificio_id:
            activos = activos.filter(ubicacion__piso__edificio_id=edificio_id)
        if ubicacion_nombre:
            activos = activos.filter(ubicacion__nombre__icontains=ubicacion_nombre)
        if tipo_uso:
            activos = activos.filter(tipo_uso=tipo_uso)
        if tipo_red:
            activos = activos.filter(tipo_red=tipo_red)
        if search_query:
            activos = activos.filter(
                Q(numero_serie__icontains=search_query) |
                Q(etiqueta__icontains=search_query) |
                Q(bdo__icontains=search_query) |
                Q(netbios__icontains=search_query) |
                Q(asignado_a__nombre__icontains=search_query) |
                Q(catalogo__modelo__icontains=search_query) |
                Q(catalogo__marca__nombre__icontains=search_query) |
                Q(catalogo__categoria__nombre__icontains=search_query)
            )

        # 2. Preparamos los datos en una lista de diccionarios
        data = []
        for activo in activos:
            data.append({
                'CATEGORIA': activo.catalogo.categoria.nombre if activo.catalogo and activo.catalogo.categoria else '',
                'MARCA': activo.catalogo.marca.nombre if activo.catalogo and activo.catalogo.marca else '',
                'MODELO': activo.catalogo.modelo if activo.catalogo else '',
                'NUMERO_SERIE': activo.numero_serie or '',
                'ETIQUETA': activo.etiqueta or '',
                'BDO': activo.bdo or '',
                'NETBIOS': activo.netbios or '',
                'TIPO_RED': activo.get_tipo_red_display(),
                'TIPO_USO': activo.get_tipo_uso_display(),
                'ESTADO': activo.estado.nombre if activo.estado else '',
                'EDIFICIO': activo.ubicacion.piso.edificio.nombre if activo.ubicacion and activo.ubicacion.piso else '',
                'PISO': activo.ubicacion.piso.nombre if activo.ubicacion else '',
                'UBICACION': activo.ubicacion.nombre if activo.ubicacion else '',
                'ASIGNATARIO': activo.asignado_a.nombre if activo.asignado_a else '',
                'CARGO_ASIGNATARIO': activo.asignado_a.cargo.nombre if activo.asignado_a and activo.asignado_a.cargo else '',
                'FECHA_REGISTRO': activo.created_at.strftime("%d/%m/%Y %H:%M") if activo.created_at else '',
                'ULTIMA_MODIFICACION': activo.updated_at.strftime("%d/%m/%Y %H:%M") if activo.updated_at else '',
            })

        # 3. Convertimos a un DataFrame de Pandas
        df = pd.DataFrame(data)

        # 4. Preparamos la respuesta HTTP como un archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # --- GENERACIÓN DINÁMICA DEL NOMBRE DEL ARCHIVO ---
        elementos_nombre = []
        
        # Identificar la categoría principal
        if categoria_nombre:
            elementos_nombre.append(categoria_nombre)
        elif categoria_id:
            try:
                elementos_nombre.append(Categoria.objects.get(id=categoria_id).nombre)
            except Categoria.DoesNotExist:
                pass
                
        # Si no hay categoría, indicar que es un reporte general
        if not elementos_nombre:
            elementos_nombre.append("General")
            
        # Añadir contexto espacial o lógico
        if ubicacion_nombre:
            elementos_nombre.append(ubicacion_nombre)
            
        if tipo_red == 'ISLA':
            elementos_nombre.append("Islas")
            
        if tipo_uso == 'EVE':
            elementos_nombre.append("Eventos")
            
        # Añadir estado si se está filtrando por él
        if estado_id:
            try:
                elementos_nombre.append(Estado.objects.get(id=estado_id).nombre)
            except Estado.DoesNotExist:
                pass
                
        # Indicar si hubo texto de búsqueda
        if search_query:
            elementos_nombre.append("Buscados")

        # Unir todos los elementos, reemplazar espacios por guiones bajos y limpiar
        nombre_crudo = "_".join(elementos_nombre).replace(" ", "_").replace("/", "_")
        
        # Evitar guiones bajos dobles (ej. "General__Operativo")
        while "__" in nombre_crudo:
            nombre_crudo = nombre_crudo.replace("__", "_")
            
        nombre_archivo = f"Inventario_{nombre_crudo}.xlsx"
        
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

        # 5. Usamos el motor de openpyxl de Pandas para escribir el archivo directamente en la respuesta
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            # Escribimos los datos
            df.to_excel(writer, index=False, sheet_name='Inventario')
            
            # Opcional: Damos un poco de formato a la cabecera
            worksheet = writer.sheets['Inventario']
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")
                
            # Ajustamos el ancho de las columnas automáticamente basado en la cabecera
            for idx, col in enumerate(df.columns):
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = max(len(col) + 2, 15)

        return response




class AuditoriaListView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = AuditoriaActivo
    template_name = 'auditoria_lista.html'
    context_object_name = 'registros'
    paginate_by = 30
    group_required = ['ADR', 'Operador ADR']

    def get_queryset(self):
        queryset = super().get_queryset().select_related('usuario', 'content_type')
        
        # Filtros básicos
        usuario = self.request.GET.get('usuario')
        if usuario and usuario != 'todos':
            queryset = queryset.filter(usuario__username=usuario)
            
        accion = self.request.GET.get('accion')
        if accion and accion != 'todas':
            queryset = queryset.filter(accion=accion)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuarios'] = User.objects.filter(is_active=True)
        context['acciones'] = AuditoriaActivo.TipoAccion.choices
        return context




class ActivosEliminadosListView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = Activo
    template_name = 'lista_eliminados.html'
    context_object_name = 'activos'
    paginate_by = 15
    group_required = ['ADR', 'Alumno en Práctica', 'Auxiliar Operador ADR', 'Operador ADR']

    def get_queryset(self):
        # Filtramos solo los que tienen el soft-delete activo
        queryset = Activo.all_objects.filter(is_deleted=True).select_related(
            'catalogo__categoria', 'catalogo__marca', 'ubicacion'
        )
        
        # Mantenemos la lógica de búsqueda que ya tienes en lista_activos.html
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(bdo__icontains=q) | 
                Q(numero_serie__icontains=q) | 
                Q(etiqueta__icontains=q)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_lista'] = "Papelera de Activos (Eliminados)"
        return context




class RestaurarActivoView(LoginRequiredMixin, GroupRequiredMixin, View):
    """
    Vista para restaurar un activo eliminado (Soft Delete).
    Cambia el estado de is_deleted a False.
    """

    group_required = ['ADR', 'Operador ADR']

    def post(self, request, pk):
        # Buscamos en all_objects porque el manager principal filtra los eliminados
        activo = get_object_or_404(Activo.all_objects, pk=pk)
        
        activo.is_deleted = False
        activo.save() # Esto dispara automáticamente las señales de auditoría
        
        messages.success(
            request, 
            f"El equipo {activo.catalogo} ({activo.numero_serie or activo.etiqueta}) ha sido restaurado."
        )
        return redirect('lista_eliminados')
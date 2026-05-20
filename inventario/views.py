from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse, reverse_lazy
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q, Count, ProtectedError
from django.views.generic import TemplateView, ListView, UpdateView, DetailView, CreateView, DeleteView, View
from django.http import HttpResponse
from django.db import transaction
from django.core.exceptions import ValidationError
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from django.contrib.auth.models import User
from django.contrib.contenttypes.models import ContentType
from django.core.cache import cache


from accounts.mixins import GroupRequiredMixin
from .forms import ActivoForm, CatalogoForm, CategoriaForm, AreaAdministrativaForm, CargoForm, FuncionarioForm
from .utils import _get_excel_val
from .models import Activo, Edificio, Piso, Ubicacion, Marca, Categoria, Estado, Catalogo, Funcionario, AuditoriaActivo, AreaAdministrativa, Cargo
from adr.models import Prestamo


class DashboardInventario(LoginRequiredMixin, GroupRequiredMixin, TemplateView):
    template_name = 'inventario/pages/dashboard.html'
    group_required = ['ADR', 'Alumno en Práctica', 'Auxiliar Operador ADR', 'Operador ADR']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Filtro base para activos no eliminados
        activos_qs = Activo.objects.filter(is_deleted=False) #
        
        # KPIs Principales
        context['total_activos'] = activos_qs.count()
        context['activos_azotea'] = activos_qs.filter(ubicacion__nombre='Azotea').count()
        context['activos_bodega'] = activos_qs.filter(ubicacion__nombre='Bodega ADR').count()
        
        # Gestión de Préstamos
        prestamos_activos = Prestamo.objects.filter(estado='En Préstamo') #
        context['prestamos_count'] = prestamos_activos.count()
        context['prestamos_pendientes'] = prestamos_activos.order_by('-fecha_prestamo')[:6]

        # Categorías con Caché (15 minutos)
        categorias = cache.get('dash_categorias')
        if not categorias:
            categorias = Categoria.objects.all().order_by('nombre') #
            cache.set('dash_categorias', categorias, 900)
        context['categorias'] = categorias

        # Ubicaciones de interés para ADR
        nombres_adr = ["Oficina ADR", "Bodega ADR", "Bodega Patio", "Azotea", "Bodega Central", "Auditorio"]
        context['ubicaciones_clave'] = Ubicacion.objects.filter(nombre__in=nombres_adr) #

        # Historial de Auditoría
        context['auditoria'] = AuditoriaActivo.objects.select_related('usuario', 'content_type').all()[:8] #

        return context




class ListaActivosView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    """
    Vista unificada para listar activos con filtros avanzados por Query Params.
    Reemplaza a activo_list.
    """
    model = Activo
    template_name = 'inventario/pages/lista_activos.html'
    paginate_by = 30
    context_object_name = 'activos'
    group_required = ['ADR', 'Alumno en Práctica', 'Auxiliar Operador ADR', 'Operador ADR']
    
    def get_queryset(self):
        # 1. Optimización de la consulta con select_related
        queryset = Activo.objects.select_related(
            'catalogo__categoria', 'catalogo__marca', 
            'ubicacion__piso__edificio', 'asignado_a', 'estado', 'creado_por',
            'asignado_a__cargo', 'asignado_a__area'
        ).all()
        
        # 2. Captura de parámetros GET (con strip para limpiar espacios en blanco fantasma)
        self.search_query = self.request.GET.get('search', '').strip()
        self.categoria_id = self.request.GET.get('categoria', '')
        self.categoria_nombre = self.request.GET.get('categoria_nombre', '')
        self.marca_id = self.request.GET.get('marca', '')
        self.modelo_str = self.request.GET.get('modelo', '')
        self.estado_id = self.request.GET.get('estado', '')
        self.edificio_id = self.request.GET.get('edificio', '')
        self.piso_id = self.request.GET.get('piso', '')
        self.ubicacion_id = self.request.GET.get('ubicacion', '')
        self.ubicacion_nombre = self.request.GET.get('ubicacion_nombre', '')
        self.tipo_uso = self.request.GET.get('tipo_uso', '')
        self.tipo_red = self.request.GET.get('tipo_red', '')
        self.asignatario_id = self.request.GET.get('asignatario', '')
        self.fecha_desde = self.request.GET.get('fecha_desde', '')
        self.fecha_hasta = self.request.GET.get('fecha_hasta', '')
        self.orden = self.request.GET.get('orden', '-updated_at') # Orden por defecto

        # Título base por defecto
        self.titulo_lista = "Listado General de Activos"

        # 3. Aplicación dinámica de filtros
        # --- CATEGORÍAS ---
        if self.categoria_id:
            queryset = queryset.filter(catalogo__categoria_id=self.categoria_id)
            try:
                cat_obj = Categoria.objects.get(id=self.categoria_id)
                self.titulo_lista = f"Inventario de {cat_obj.nombre}s"
            except Categoria.DoesNotExist:
                pass

        elif self.categoria_nombre:
            queryset = queryset.filter(catalogo__categoria__nombre__icontains=self.categoria_nombre)
            self.titulo_lista = f"Inventario de {self.categoria_nombre}s"
            cat_obj = Categoria.objects.filter(nombre__iexact=self.categoria_nombre).first()
            if cat_obj:
                self.categoria_id = str(cat_obj.id)


        if self.marca_id:
            queryset = queryset.filter(catalogo__marca_id=self.marca_id)
            
        if self.modelo_str:
            queryset = queryset.filter(catalogo__modelo=self.modelo_str)

        if self.estado_id:
            queryset = queryset.filter(estado_id=self.estado_id)

        if self.tipo_uso:
            queryset = queryset.filter(tipo_uso=self.tipo_uso)
            if self.tipo_uso == Activo.TipoUso.EVENTOS:
                self.titulo_lista = "Equipos para Eventos"

        if self.tipo_red:
            queryset = queryset.filter(tipo_red=self.tipo_red)
            if self.tipo_red == Activo.TipoRed.ISLA:
                self.titulo_lista = "Equipos Isla"
                
        if self.asignatario_id:
            queryset = queryset.filter(asignado_a_id=self.asignatario_id)
            
        if self.fecha_desde:
            queryset = queryset.filter(created_at__date__gte=self.fecha_desde)
            
        if self.fecha_hasta:
            queryset = queryset.filter(created_at__date__lte=self.fecha_hasta)
            
        # --- UBICACIONES ---
        ub_obj = None
            
        if self.ubicacion_id:
            queryset = queryset.filter(ubicacion_id=self.ubicacion_id)
            ub_obj = Ubicacion.objects.filter(id=self.ubicacion_id).select_related('piso__edificio').first()
        elif self.ubicacion_nombre:
            queryset = queryset.filter(ubicacion__nombre__icontains=self.ubicacion_nombre)
            ub_obj = Ubicacion.objects.filter(nombre__iexact=self.ubicacion_nombre).select_related('piso__edificio').first()
            if ub_obj:
                # Sincronizamos el ID para que el selector del HTML se marque
                self.ubicacion_id = str(ub_obj.id)
        
        # Si hay una ubicación (ya sea por ID o Nombre), actualizamos título y jerarquía
        if ub_obj:
            suffix = f" en {ub_obj.nombre}"
            # Si el título es el general, usamos "Inventario en...", si no, concatenamos
            if self.titulo_lista == "Listado General de Activos":
                self.titulo_lista = f"Inventario{suffix}"
            else:
                self.titulo_lista += suffix
            
            self.piso_id = str(ub_obj.piso_id)
            self.edificio_id = str(ub_obj.piso.edificio_id)

        # Filtros de Piso o Edificio si no se seleccionó una ubicación específica
        elif self.piso_id:
            pi_obj = Piso.objects.filter(id=self.piso_id).first()
            if pi_obj:
                suffix = f" en {pi_obj.nombre}"
                if self.titulo_lista == "Listado General de Activos":
                    self.titulo_lista = f"Inventario{suffix}"
                else:
                    self.titulo_lista += suffix
                self.edificio_id = str(pi_obj.edificio_id)
        
        elif self.edificio_id:
            ed_obj = Edificio.objects.filter(id=self.edificio_id).first()
            if ed_obj:
                suffix = f" en {ed_obj.nombre}"
                if self.titulo_lista == "Listado General de Activos":
                    self.titulo_lista = f"Inventario{suffix}"
                else:
                    self.titulo_lista += suffix


        # 4. Búsqueda de texto global (Ampliando la capacidad del buscador)
        if self.search_query:
            queryset = queryset.filter(
                Q(numero_serie__icontains=self.search_query) |
                Q(etiqueta__icontains=self.search_query) |
                Q(bdo__icontains=self.search_query) |
                Q(netbios__icontains=self.search_query) |
                Q(asignado_a__nombre__icontains=self.search_query) |
                Q(catalogo__modelo__icontains=self.search_query) |
                Q(catalogo__marca__nombre__icontains=self.search_query) |
                Q(catalogo__categoria__nombre__icontains=self.search_query)
            )

        # 5. Aplicar ordenamiento
        valid_orders = ['created_at', '-created_at', 'updated_at', '-updated_at']
        if self.orden in valid_orders:
            queryset = queryset.order_by(self.orden)
        else:
            queryset = queryset.order_by('-updated_at')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Preservar query params para la paginación
        get_copy = self.request.GET.copy()
        if 'page' in get_copy:
            get_copy.pop('page')
        query_string = get_copy.urlencode()

        context.update({
            'titulo_lista': getattr(self, 'titulo_lista', "Listado General de Activos"),
            'query_string': f"&{query_string}" if query_string else "",
            
            # Listas para rellenar los select del HTML
            'categorias': Categoria.objects.all().order_by('nombre'),
            'marcas': Marca.objects.all().order_by('nombre'),
            'modelos': Catalogo.objects.values_list('modelo', flat=True).distinct().order_by('modelo'),
            'estados': Estado.objects.all().order_by('nombre'),
            'edificios': Edificio.objects.all().order_by('nombre'),
            'pisos': Piso.objects.select_related('edificio').all().order_by('edificio__nombre', 'nombre'),
            'ubicaciones': Ubicacion.objects.select_related('piso__edificio').all().order_by('piso__edificio__nombre', 'piso__nombre', 'nombre'),
            'asignatarios': Funcionario.objects.all().order_by('nombre'),
            'tipos_uso': Activo.TipoUso.choices,
            'tipos_red': Activo.TipoRed.choices,
            
            # Variables para mantener seleccionada la opción correcta en la vista
            'search_query': getattr(self, 'search_query', ''),
            'categoria_seleccionada': getattr(self, 'categoria_id', ''),
            'marca_seleccionada': getattr(self, 'marca_id', ''),
            'modelo_seleccionado': getattr(self, 'modelo_str', ''),
            'estado_seleccionado': getattr(self, 'estado_id', ''),
            'edificio_seleccionado': getattr(self, 'edificio_id', ''),
            'piso_seleccionado': getattr(self, 'piso_id', ''),
            'ubicacion_seleccionada': getattr(self, 'ubicacion_id', ''),
            'tipo_uso_seleccionado': getattr(self, 'tipo_uso', ''),
            'tipo_red_seleccionado': getattr(self, 'tipo_red', ''),
            'asignatario_seleccionado': getattr(self, 'asignatario_id', ''),
            'fecha_desde': getattr(self, 'fecha_desde', ''),
            'fecha_hasta': getattr(self, 'fecha_hasta', ''),
            'orden_seleccionado': getattr(self, 'orden', '-updated_at'),
        })
        return context




class EditarActivoView(LoginRequiredMixin, GroupRequiredMixin, UpdateView):
    """
    Vista unificada para editar cualquier tipo de activo.
    """
    model = Activo
    form_class = ActivoForm
    template_name = 'inventario/pages/editar_activo.html'
    context_object_name = 'activo'
    group_required = ['ADR', 'Auxiliar Operador ADR', 'Operador ADR']
    success_url = reverse_lazy('lista_activos')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # self.object es la instancia del activo que estamos editando
        context['titulo_form'] = f"Editar {self.object.catalogo.categoria.nombre}"
        return context

    def form_valid(self, form):
        # Dispara el guardado y emite el mensaje de éxito
        response = super().form_valid(form)
        messages.success(self.request, f'El activo {self.object} ha sido actualizado correctamente.')
        return response

    def form_invalid(self, form):
        # Emite el mensaje de error si la validación falla
        messages.error(self.request, 'No se pudo guardar. Por favor, corrige los errores en el formulario.')
        return super().form_invalid(form)

    def get_success_url(self):
        # Redirección dinámica
        params = self.request.GET.urlencode()
        url = self.success_url
        if params:
            return f"{url}?{params}"
        return url




class DetalleActivoView(LoginRequiredMixin, GroupRequiredMixin, DetailView):
    """
    Vista para mostrar el detalle completo de un activo.
    """
    model = Activo
    template_name = 'inventario/pages/ver_activo.html'
    context_object_name = 'activo'
    group_required = ['ADR', 'Alumno en Práctica', 'Auxiliar Operador ADR', 'Operador ADR']

    def get_queryset(self):
        # Pre-cargamos todas las relaciones para que la vista sea rápida y eficiente
        return Activo.objects.select_related(
            'catalogo__categoria', 'catalogo__marca', 
            'estado', 'ubicacion__piso__edificio', 
            'asignado_a__cargo', 'asignado_a__area'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_detalle'] = f"Detalle del Equipo: {self.object.catalogo.marca} {self.object.catalogo.modelo or ''}".strip()
        return context




class AgregarActivoView(LoginRequiredMixin, GroupRequiredMixin, CreateView):
    model = Activo
    form_class = ActivoForm
    template_name = 'inventario/pages/agregar_activo.html'
    group_required = ['ADR', 'Auxiliar Operador ADR', 'Operador ADR']
    success_url = reverse_lazy('lista_activos')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if 'catalogo_form' not in context:
            context['catalogo_form'] = CatalogoForm()
        return context

    def post(self, request, *args, **kwargs):
        self.object = None
        form = self.get_form()
        crear_nuevo = request.POST.get('crear_nuevo_catalogo') == 'true'
        
        if crear_nuevo:
            with transaction.atomic():
                # El form procesa la Marca automáticamente en su clean_marca
                catalogo_form = CatalogoForm(request.POST, request.FILES)
                if catalogo_form.is_valid():
                    nuevo_catalogo = catalogo_form.save()
                    
                    ## Log de Auditoría para el nuevo Catálogo
                    #AuditoriaActivo.objects.create(
                    #    usuario=request.user,
                    #    accion=AuditoriaActivo.TipoAccion.CREACION,
                    #    content_type=ContentType.objects.get_for_model(Catalogo),
                    #    object_id=nuevo_catalogo.id,
                    #    valor_nuevo=f"Producto creado desde activo: {nuevo_catalogo}"
                    #)
                    
                    data_activo = request.POST.copy()
                    data_activo['catalogo'] = nuevo_catalogo.id
                    form = ActivoForm(data_activo)
                else:
                    return self.render_to_response(self.get_context_data(form=form, catalogo_form=catalogo_form))

        if form.is_valid():
            return self.form_valid(form)
        return self.form_invalid(form)

    def form_valid(self, form):
        # 1. Asignamos el usuario autenticado
        form.instance.creado_por = self.request.user
        response = super().form_valid(form)
        
        ## 2. Log de Auditoría para el Activo (EL REGISTRO DE AUDITORIA SE CREA EN SIGNALS. NO HACE FATA QUE ESTÉ AQUÍ)
        #AuditoriaActivo.objects.create(
        #    usuario=self.request.user,
        #    accion=AuditoriaActivo.TipoAccion.CREACION,
        #    content_type=ContentType.objects.get_for_model(Activo),
        #    object_id=self.object.id,
        #    valor_nuevo=f"Activo registrado: {self.object}"
        #)
        messages.success(self.request, "Activo y auditoría registrados correctamente.")
        return response


class EliminarActivoView(LoginRequiredMixin, GroupRequiredMixin, DeleteView):
    """
    Vista para procesar la eliminación (soft-delete) de un activo.
    """
    model = Activo
    template_name = 'inventario/pages/eliminar_activo.html'
    context_object_name = 'activo'
    success_url = reverse_lazy('lista_activos')
    group_required = ['ADR', 'Operador ADR']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_form'] = 'Confirmar Eliminación'
        return context

    def form_valid(self, form):
        # En Django moderno (4.0+), form_valid maneja la lógica de DeleteView
        activo = self.get_object()
        # Llamar a delete() invoca override en adr/models.py (soft-delete)
        activo.delete()
        messages.success(self.request, f'El equipo {activo} ha sido enviado a la papelera (Eliminado).')
        return redirect(self.get_success_url())
    
    def get_success_url(self):
        # Redirección dinámica
        params = self.request.GET.urlencode()
        url = self.success_url
        if params:
            return f"{url}?{params}"
        return url




class EliminarActivosMasivoView(LoginRequiredMixin, GroupRequiredMixin, View):
    """
    Vista para procesar la eliminación masiva (soft-delete) mediante checkboxes.
    """
    group_required = ['ADR', 'Operador ADR']

    def post(self, request, *args, **kwargs):
        # 'activos_seleccionados' será el atributo 'name' de nuestros checkboxes en el HTML
        activos_ids = request.POST.getlist('activos_seleccionados')
        
        if not activos_ids:
            messages.warning(request, "No se seleccionó ningún equipo para eliminar.")
            return redirect('lista_activos')

        try:
            # Filtramos los activos que coincidan con los IDs recibidos
            activos_a_eliminar = Activo.objects.filter(id__in=activos_ids)
            cantidad = activos_a_eliminar.count()

            # Iteramos para asegurarnos de que se ejecute el método delete() personalizado
            # y cualquier señal o registro de auditoría acoplado al guardado.
            for activo in activos_a_eliminar:
                activo.delete() # Esto ejecuta tu soft-delete (is_deleted = True)

            messages.success(request, f'Se han enviado {cantidad} equipos a la papelera correctamente.')
        
        except Exception as e:
            messages.error(request, f'Ocurrió un error durante la eliminación masiva: {str(e)}')

        # Redirección dinámica
        params = request.GET.urlencode()
        url = reverse('lista_activos')
        if params:
            return redirect(f"{url}?{params}")
        return redirect(url)




class SubirExcelActivosView(LoginRequiredMixin, GroupRequiredMixin, View):
    """
    Vista para importar activos masivamente mediante Excel.
    Aplica reglas estrictas, mapea etiquetas legibles y registra al usuario en auditoría.
    """
    template_name = 'inventario/pages/subir_excel_activos.html'
    group_required = ['ADR', 'Auxiliar Operador ADR', 'Operador ADR']

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        if 'archivo_excel' not in request.FILES:
            messages.error(request, 'Por favor, selecciona un archivo Excel válido.')
            return redirect('subir_excel_activos')
            
        archivo = request.FILES['archivo_excel']
        
        if not archivo.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'El formato del archivo no es válido. Debe ser .xlsx o .xls')
            return redirect('subir_excel_activos')
            
        try:
            # sheet_name=0 asegura que siempre lea la primera hoja ("Equipos")
            df = pd.read_excel(archivo, sheet_name=0)
            
            # Limpiamos nombres de columnas (quitar espacios y poner a mayúsculas)
            df.columns = df.columns.str.strip().str.upper()
            
            # Reemplazar valores NaN por None
            df = df.where(pd.notnull(df), None)

            # MAPEO INVERSO
            red_map = {str(label).strip().upper(): key for key, label in Activo.TipoRed.choices}
            uso_map = {str(label).strip().upper(): key for key, label in Activo.TipoUso.choices}
            
            registros_exitosos = 0
            errores = []

            with transaction.atomic():
                for index, row in df.iterrows():
                    try:
                        # 1. Validar Categoría
                        cat_nombre = _get_excel_val(row, 'CATEGORIA', default='', to_upper=True)
                        if not cat_nombre:
                            errores.append(f"Fila {index + 2}: La categoría es obligatoria.")
                            continue
                            
                        try:
                            categoria = Categoria.objects.get(nombre__iexact=cat_nombre)
                        except Categoria.DoesNotExist:
                            errores.append(f"Fila {index + 2}: La categoría '{cat_nombre}' no existe en el sistema. Registro denegado.")
                            continue

                        # 2. Validar Ubicación
                        ubicacion_nombre = _get_excel_val(row, 'UBICACION', default='')
                        ubicacion_obj = None
                        if ubicacion_nombre:
                            nombre_edi = _get_excel_val(row, 'EDIFICIO', default='')
                            nombre_piso = _get_excel_val(row, 'PISO', default='')

                            if not nombre_edi or not nombre_piso:
                                errores.append(f"Fila {index + 2}: Para asignar la ubicación '{ubicacion_nombre}', debe especificar EDIFICIO y PISO.")
                                continue

                            ubicacion_obj = Ubicacion.objects.filter(
                                nombre__iexact=ubicacion_nombre,
                                piso__nombre__iexact=nombre_piso,
                                piso__edificio__nombre__iexact=nombre_edi
                            ).first()
                            
                            if not ubicacion_obj:
                                errores.append(f"Fila {index + 2}: La ubicación '{ubicacion_nombre}' (Piso: {nombre_piso}, Edificio: {nombre_edi}) no existe en el sistema. Registro denegado.")
                                continue

                        # 3. Procesar Marca y Modelo
                        marca_nombre = _get_excel_val(row, 'MARCA', default='', to_upper=True)
                        if not marca_nombre:
                            errores.append(f"Fila {index + 2}: La marca es obligatoria.")
                            continue
                            
                        modelo_nombre = _get_excel_val(row, 'MODELO', default='GENÉRICO', to_upper=True)
                        
                        marca, _ = Marca.objects.get_or_create(nombre=marca_nombre)
                        catalogo, _ = Catalogo.objects.get_or_create(
                            categoria=categoria, 
                            marca=marca, 
                            modelo=modelo_nombre
                        )

                        # 4. Procesar Estado y Asignatario
                        estado_nombre = _get_excel_val(row, 'ESTADO', default='OPERATIVO', to_upper=True)
                        estado, _ = Estado.objects.get_or_create(nombre=estado_nombre)

                        funcionario_nombre = _get_excel_val(row, 'ASIGNATARIO', default='', to_upper=True)
                        funcionario_obj = None
                        if funcionario_nombre:
                            funcionario_obj, _ = Funcionario.objects.get_or_create(nombre=funcionario_nombre)

                        # 5. Mapeo de Red y Uso
                        tipo_red_label = _get_excel_val(row, 'TIPO_RED', default='', to_upper=True)
                        tipo_uso_label = _get_excel_val(row, 'TIPO_USO', default='', to_upper=True)

                        tipo_red = red_map.get(tipo_red_label, 'DOM')
                        tipo_uso = uso_map.get(tipo_uso_label, 'PER')

                        # 6. Creación/Actualización del Activo
                        numero_serie = _get_excel_val(row, 'NUMERO_SERIE', default=None)
                        etiqueta = _get_excel_val(row, 'ETIQUETA', default=None)
                        bdo = _get_excel_val(row, 'BDO', default=None)
                        netbios = _get_excel_val(row, 'NETBIOS', default=None)

                        activo = None
                        if bdo:
                            activo = Activo.objects.filter(bdo=bdo).first()
                        if not activo and numero_serie:
                            activo = Activo.objects.filter(numero_serie=numero_serie).first()

                        # Obtenemos el ContentType del modelo Activo para la Auditoría
                        activo_ct = ContentType.objects.get_for_model(Activo)

                        if activo:
                            # --- MODO EDICIÓN ---
                            activo.catalogo = catalogo
                            activo.estado = estado
                            activo.ubicacion = ubicacion_obj
                            activo.asignado_a = funcionario_obj
                            activo.tipo_red = tipo_red
                            activo.tipo_uso = tipo_uso
                            if netbios: activo.netbios = netbios
                            
                            activo.save()

                            # REGISTRO EN TU MODELO DE AUDITORÍA (MODIFICACIÓN)
                            AuditoriaActivo.objects.create(
                                usuario=request.user,
                                accion=AuditoriaActivo.TipoAccion.MODIFICACION,
                                content_type=activo_ct,
                                object_id=activo.id,
                                campo="Múltiples (Importación Excel)",
                                valor_nuevo="Actualizado vía Excel"
                            )
                        else:
                            # --- MODO CREACIÓN ---
                            nuevo_activo = Activo(
                                catalogo=catalogo,
                                estado=estado,
                                numero_serie=numero_serie,
                                etiqueta=etiqueta,
                                bdo=bdo,
                                netbios=netbios,
                                tipo_red=tipo_red,
                                tipo_uso=tipo_uso,
                                ubicacion=ubicacion_obj,
                                asignado_a=funcionario_obj,
                                creado_por=request.user  # ASIGNAMOS EL CREADOR REAL AQUÍ
                            )
                            
                            nuevo_activo.save()
                            
                            # REGISTRO EN TU MODELO DE AUDITORÍA (CREACIÓN)
                            AuditoriaActivo.objects.create(
                                usuario=request.user,
                                accion=AuditoriaActivo.TipoAccion.CREACION,
                                content_type=activo_ct,
                                object_id=nuevo_activo.id,
                                campo="Todos (Importación Excel)",
                                valor_nuevo="Creado masivamente vía Excel"
                            )
                            
                        registros_exitosos += 1

                    # 1. Captura ESPECÍFICA para validaciones del modelo
                    except ValidationError as e:
                        # Extraemos los mensajes de forma limpia (sin diccionarios crudos)
                        mensajes_error = ", ".join(e.messages)
                        errores.append(f"Fila {index + 2}: {mensajes_error}")
                        
                    # 2. Captura genérica para otros fallos (ej. base de datos)
                    except Exception as e:
                        errores.append(f"Fila {index + 2}: Error interno ({str(e)})")

            if errores:
                # Si hay errores, renderiza la misma plantilla con el detalle.
                context = {
                    'errores': errores,
                    'registros_exitosos': registros_exitosos,
                }
                messages.warning(request, f'Se importaron {registros_exitosos} equipos, pero hubo {len(errores)} errores. Revisa el detalle abajo.')
                return render(request, self.template_name, context)

            if registros_exitosos > 0:
                messages.success(request, f'Proceso completado. Se importaron/actualizaron {registros_exitosos} activos.')
            else:
                messages.info(request, 'El proceso finalizó, pero no se importó ningún registro nuevo.')
                
            # Retorno unificado a la lista principal
            return redirect('lista_activos')

        except Exception as e:
            messages.error(request, f'Error al leer el archivo Excel: {str(e)}')
            return redirect('subir_excel_activos')




class DescargarPlantillaExcelView(LoginRequiredMixin, GroupRequiredMixin, View):
    """
    Genera un archivo Excel (.xlsx) con las cabeceras correctas y listas desplegables
    basadas en los datos actuales del sistema (incluyendo etiquetas legibles para choices).
    """
    group_required = ['ADR', 'Alumno en Práctica', 'Auxiliar Operador ADR', 'Operador ADR']

    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        
        # --- HOJA 1: EQUIPOS ---
        ws_equipos = wb.active
        ws_equipos.title = "Equipos"
        
        cabeceras = [
            'CATEGORIA', 'MARCA', 'MODELO', 'NUMERO_SERIE', 'ETIQUETA', 
            'BDO', 'NETBIOS', 'TIPO_RED', 'TIPO_USO', 'ESTADO', 
            'EDIFICIO', 'PISO', 'UBICACION', 'ASIGNATARIO'
        ]
        
        header_fill = PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, header_title in enumerate(cabeceras, 1):
            cell = ws_equipos.cell(row=1, column=col_num, value=header_title)
            cell.fill = header_fill
            cell.font = header_font
            ws_equipos.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

        # --- HOJA 2: DATOS ---
        ws_datos = wb.create_sheet(title="Datos")
        
        # 1. Obtener datos de la BD
        categorias = list(Categoria.objects.values_list('nombre', flat=True).order_by('nombre'))
        estados = list(Estado.objects.values_list('nombre', flat=True).order_by('nombre'))
        edificios = list(Edificio.objects.values_list('nombre', flat=True).order_by('nombre'))
        pisos = list(Piso.objects.values_list('nombre', flat=True).order_by('nombre'))
        ubicaciones = list(Ubicacion.objects.values_list('nombre', flat=True).order_by('nombre'))
        marcas = list(Marca.objects.values_list('nombre', flat=True).order_by('nombre'))
        
        # IMPORTANTE: Ahora extraemos index 1 (Label humano) en lugar de index 0 (Código de BD)
        tipos_red_labels = [choice[1] for choice in Activo.TipoRed.choices]
        tipos_uso_labels = [choice[1] for choice in Activo.TipoUso.choices]

        # 2. Escribir datos
        datos_diccionario = {
            1: categorias,        # Columna A
            2: estados,           # Columna B
            3: tipos_red_labels,  # Columna C (Humanos)
            4: tipos_uso_labels,  # Columna D (Humanos)
            5: edificios,         # Columna E
            6: pisos,             # Columna F
            7: ubicaciones,       # Columna G
            8: marcas             # Columna H (NUEVO: Marcas)
        }

        for col_index, valores in datos_diccionario.items():
            for row_index, valor in enumerate(valores, 1):
                ws_datos.cell(row=row_index, column=col_index, value=valor)

        ws_datos.sheet_state = 'hidden'

        # --- APLICAR VALIDACIÓN DE DATOS (Desplegables) ---
        def crear_dropdown(col_letra_datos, total_filas, col_letra_equipos, allow_blank=True):
            if total_filas > 0:
                formula = f"Datos!${col_letra_datos}$1:${col_letra_datos}${total_filas}"
                # Para marcas, permitimos espacios en blanco/escribir advertencias
                dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank, showErrorMessage=False)
                dv.add(f"{col_letra_equipos}2:{col_letra_equipos}1000")
                ws_equipos.add_data_validation(dv)

        crear_dropdown('A', len(categorias), 'A')        # A = CATEGORIA
        crear_dropdown('B', len(estados), 'J')            # J = ESTADO
        crear_dropdown('C', len(tipos_red_labels), 'H')   # H = TIPO_RED
        crear_dropdown('D', len(tipos_uso_labels), 'I')   # I = TIPO_USO
        crear_dropdown('E', len(edificios), 'K')          # K = EDIFICIO
        crear_dropdown('F', len(pisos), 'L')              # L = PISO
        crear_dropdown('G', len(ubicaciones), 'M')        # M = UBICACION
        crear_dropdown('H', len(marcas), 'B')             # B = MARCA (NUEVO)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Plantilla_Ingreso_Activos.xlsx"'
        
        wb.save(response)
        return response




class DescargarExcelActivosView(LoginRequiredMixin, GroupRequiredMixin, View):
    """
    Vista para exportar los activos registrados en el sistema a un archivo Excel.
    Aplica de manera dinámica los mismos filtros que la vista de listado.
    """
    group_required = ['ADR', 'Alumno en Práctica', 'Auxiliar Operador ADR', 'Operador ADR']

    def get(self, request, *args, **kwargs):
        # 1. Obtenemos todos los activos optimizando las relaciones
        activos = Activo.objects.select_related(
            'catalogo__categoria', 'catalogo__marca', 'estado',
            'ubicacion__piso__edificio', 'asignado_a__cargo'
        ).all().order_by('-updated_at')

        # Capturamos los mismos Query Params que usa la lista de activos
        search_query = request.GET.get('search', '')
        categoria_id = request.GET.get('categoria', '')
        categoria_nombre = request.GET.get('categoria_nombre', '')
        marca_id = request.GET.get('marca', '')
        estado_id = request.GET.get('estado', '')
        edificio_id = request.GET.get('edificio', '')
        ubicacion_nombre = request.GET.get('ubicacion_nombre', '')
        tipo_uso = request.GET.get('tipo_uso', '')
        tipo_red = request.GET.get('tipo_red', '')

        # Aplicamos los filtros dinámicamente si existen en la URL
        if categoria_id:
            activos = activos.filter(catalogo__categoria_id=categoria_id)
        if categoria_nombre:
            activos = activos.filter(catalogo__categoria__nombre__icontains=categoria_nombre)
        if marca_id:
            activos = activos.filter(catalogo__marca_id=marca_id)
        if estado_id:
            activos = activos.filter(estado_id=estado_id)
        if edificio_id:
            activos = activos.filter(ubicacion__piso__edificio_id=edificio_id)
        if ubicacion_nombre:
            activos = activos.filter(ubicacion__nombre__icontains=ubicacion_nombre)
        if tipo_uso:
            activos = activos.filter(tipo_uso=tipo_uso)
        if tipo_red:
            activos = activos.filter(tipo_red=tipo_red)
        if search_query:
            activos = activos.filter(
                Q(numero_serie__icontains=search_query) |
                Q(etiqueta__icontains=search_query) |
                Q(bdo__icontains=search_query) |
                Q(netbios__icontains=search_query) |
                Q(asignado_a__nombre__icontains=search_query) |
                Q(catalogo__modelo__icontains=search_query) |
                Q(catalogo__marca__nombre__icontains=search_query) |
                Q(catalogo__categoria__nombre__icontains=search_query)
            )

        # 2. Preparamos los datos en una lista de diccionarios
        data = []
        for activo in activos:
            data.append({
                'CATEGORIA': activo.catalogo.categoria.nombre if activo.catalogo and activo.catalogo.categoria else '',
                'MARCA': activo.catalogo.marca.nombre if activo.catalogo and activo.catalogo.marca else '',
                'MODELO': activo.catalogo.modelo if activo.catalogo else '',
                'NUMERO_SERIE': activo.numero_serie or '',
                'ETIQUETA': activo.etiqueta or '',
                'BDO': activo.bdo or '',
                'NETBIOS': activo.netbios or '',
                'TIPO_RED': activo.get_tipo_red_display(),
                'TIPO_USO': activo.get_tipo_uso_display(),
                'ESTADO': activo.estado.nombre if activo.estado else '',
                'EDIFICIO': activo.ubicacion.piso.edificio.nombre if activo.ubicacion and activo.ubicacion.piso else '',
                'PISO': activo.ubicacion.piso.nombre if activo.ubicacion else '',
                'UBICACION': activo.ubicacion.nombre if activo.ubicacion else '',
                'ASIGNATARIO': activo.asignado_a.nombre if activo.asignado_a else '',
                'CARGO_ASIGNATARIO': activo.asignado_a.cargo.nombre if activo.asignado_a and activo.asignado_a.cargo else '',
                'FECHA_REGISTRO': activo.created_at.strftime("%d/%m/%Y %H:%M") if activo.created_at else '',
                'ULTIMA_MODIFICACION': activo.updated_at.strftime("%d/%m/%Y %H:%M") if activo.updated_at else '',
            })

        # 3. Convertimos a un DataFrame de Pandas
        df = pd.DataFrame(data)

        # 4. Preparamos la respuesta HTTP como un archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # --- GENERACIÓN DINÁMICA DEL NOMBRE DEL ARCHIVO ---
        elementos_nombre = []
        
        # Identificar la categoría principal
        if categoria_nombre:
            elementos_nombre.append(categoria_nombre)
        elif categoria_id:
            try:
                elementos_nombre.append(Categoria.objects.get(id=categoria_id).nombre)
            except Categoria.DoesNotExist:
                pass
                
        # Si no hay categoría, indicar que es un reporte general
        if not elementos_nombre:
            elementos_nombre.append("General")
            
        # Añadir contexto espacial o lógico
        if ubicacion_nombre:
            elementos_nombre.append(ubicacion_nombre)
            
        if tipo_red == 'ISLA':
            elementos_nombre.append("Islas")
            
        if tipo_uso == 'EVE':
            elementos_nombre.append("Eventos")
            
        # Añadir estado si se está filtrando por él
        if estado_id:
            try:
                elementos_nombre.append(Estado.objects.get(id=estado_id).nombre)
            except Estado.DoesNotExist:
                pass
                
        # Indicar si hubo texto de búsqueda
        if search_query:
            elementos_nombre.append("Buscados")

        # Unir todos los elementos, reemplazar espacios por guiones bajos y limpiar
        nombre_crudo = "_".join(elementos_nombre).replace(" ", "_").replace("/", "_")
        
        # Evitar guiones bajos dobles (ej. "General__Operativo")
        while "__" in nombre_crudo:
            nombre_crudo = nombre_crudo.replace("__", "_")
            
        nombre_archivo = f"Inventario_{nombre_crudo}.xlsx"
        
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

        # 5. Usamos el motor de openpyxl de Pandas para escribir el archivo directamente en la respuesta
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            # Escribimos los datos
            df.to_excel(writer, index=False, sheet_name='Inventario')
            
            # Opcional: Damos un poco de formato a la cabecera
            worksheet = writer.sheets['Inventario']
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")
                
            # Ajustamos el ancho de las columnas automáticamente basado en la cabecera
            for idx, col in enumerate(df.columns):
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = max(len(col) + 2, 15)

        return response




class DescargarExcelFiltradoView(LoginRequiredMixin, GroupRequiredMixin, View):
    """
    Vista inteligente para exportar a Excel aplicando EXACTAMENTE los mismos 
    filtros y orden que el usuario tiene activos en la vista de lista.
    """
    group_required = ['ADR', 'Alumno en Práctica', 'Auxiliar Operador ADR', 'Operador ADR']

    def get(self, request, *args, **kwargs):
        # 1. Obtenemos todos los activos optimizando las relaciones
        activos = Activo.objects.select_related(
            'catalogo__categoria', 'catalogo__marca', 'estado',
            'ubicacion__piso__edificio', 'asignado_a__cargo'
        ).all()

        # Capturamos todos los Query Params de la URL
        search_query = request.GET.get('search', '').strip()
        categoria_id = request.GET.get('categoria', '')
        marca_id = request.GET.get('marca', '')
        modelo_str = request.GET.get('modelo', '')
        estado_id = request.GET.get('estado', '')
        edificio_id = request.GET.get('edificio', '')
        piso_id = request.GET.get('piso', '')
        ubicacion_id = request.GET.get('ubicacion', '')
        tipo_uso = request.GET.get('tipo_uso', '')
        tipo_red = request.GET.get('tipo_red', '')
        asignatario_id = request.GET.get('asignatario', '')
        fecha_desde = request.GET.get('fecha_desde', '')
        fecha_hasta = request.GET.get('fecha_hasta', '')
        orden = request.GET.get('orden', '-updated_at')

        # Aplicamos los filtros idénticos a los de la lista
        if categoria_id: activos = activos.filter(catalogo__categoria_id=categoria_id)
        if marca_id: activos = activos.filter(catalogo__marca_id=marca_id)
        if modelo_str: activos = activos.filter(catalogo__modelo=modelo_str)
        if estado_id: activos = activos.filter(estado_id=estado_id)
        if edificio_id: activos = activos.filter(ubicacion__piso__edificio_id=edificio_id)
        if piso_id: activos = activos.filter(ubicacion__piso_id=piso_id)
        if ubicacion_id: activos = activos.filter(ubicacion_id=ubicacion_id)
        if tipo_uso: activos = activos.filter(tipo_uso=tipo_uso)
        if tipo_red: activos = activos.filter(tipo_red=tipo_red)
        if asignatario_id: activos = activos.filter(asignado_a_id=asignatario_id)
        if fecha_desde: activos = activos.filter(created_at__date__gte=fecha_desde)
        if fecha_hasta: activos = activos.filter(created_at__date__lte=fecha_hasta)

        if search_query:
            activos = activos.filter(
                Q(numero_serie__icontains=search_query) |
                Q(etiqueta__icontains=search_query) |
                Q(bdo__icontains=search_query) |
                Q(netbios__icontains=search_query) |
                Q(asignado_a__nombre__icontains=search_query) |
                Q(catalogo__modelo__icontains=search_query) |
                Q(catalogo__marca__nombre__icontains=search_query) |
                Q(catalogo__categoria__nombre__icontains=search_query)
            )

        # Aplicamos el ordenamiento
        valid_orders = ['created_at', '-created_at', 'updated_at', '-updated_at']
        if orden in valid_orders:
            activos = activos.order_by(orden)
        else:
            activos = activos.order_by('-updated_at')

        # 2. Preparamos los datos
        data = []
        for activo in activos:
            data.append({
                'CATEGORIA': activo.catalogo.categoria.nombre if activo.catalogo and activo.catalogo.categoria else '',
                'MARCA': activo.catalogo.marca.nombre if activo.catalogo and activo.catalogo.marca else '',
                'MODELO': activo.catalogo.modelo if activo.catalogo else '',
                'NUMERO_SERIE': activo.numero_serie or '',
                'ETIQUETA': activo.etiqueta or '',
                'BDO': activo.bdo or '',
                'NETBIOS': activo.netbios or '',
                'TIPO_RED': activo.get_tipo_red_display(),
                'TIPO_USO': activo.get_tipo_uso_display(),
                'ESTADO': activo.estado.nombre if activo.estado else '',
                'EDIFICIO': activo.ubicacion.piso.edificio.nombre if activo.ubicacion and activo.ubicacion.piso else '',
                'PISO': activo.ubicacion.piso.nombre if activo.ubicacion else '',
                'UBICACION': activo.ubicacion.nombre if activo.ubicacion else '',
                'ASIGNATARIO': activo.asignado_a.nombre if activo.asignado_a else '',
                'CARGO_ASIGNATARIO': activo.asignado_a.cargo.nombre if activo.asignado_a and activo.asignado_a.cargo else '',
                'FECHA_REGISTRO': activo.created_at.strftime("%d/%m/%Y %H:%M") if activo.created_at else '',
                'ULTIMA_MODIFICACION': activo.updated_at.strftime("%d/%m/%Y %H:%M") if activo.updated_at else '',
            })

        df = pd.DataFrame(data)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # Nombre de archivo indicando que es un reporte a medida
        response['Content-Disposition'] = 'attachment; filename="Inventario_Filtrado_A_Medida.xlsx"'

        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Inventario')
            worksheet = writer.sheets['Inventario']
            
            # Cabecera de un color distinto (Amarillo/Dorado o Verde Claro) para diferenciarlo
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="000000")
                cell.fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
                
            for idx, col in enumerate(df.columns):
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = max(len(col) + 2, 15)

        return response




class AuditoriaListView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = AuditoriaActivo
    template_name = 'inventario/pages/auditoria_lista.html'
    context_object_name = 'registros'
    paginate_by = 30
    group_required = ['ADR', 'Operador ADR']

    def get_queryset(self):
        queryset = super().get_queryset().select_related('usuario', 'content_type')
        
        # Filtros básicos
        usuario = self.request.GET.get('usuario')
        if usuario and usuario != 'todos':
            queryset = queryset.filter(usuario__username=usuario)
            
        accion = self.request.GET.get('accion')
        if accion and accion != 'todas':
            queryset = queryset.filter(accion=accion)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['usuarios'] = User.objects.filter(is_active=True)
        context['acciones'] = AuditoriaActivo.TipoAccion.choices
        return context




class ActivosEliminadosListView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = Activo
    template_name = 'inventario/pages/lista_eliminados.html'
    context_object_name = 'activos'
    paginate_by = 15
    group_required = ['ADR', 'Alumno en Práctica', 'Auxiliar Operador ADR', 'Operador ADR']

    def get_queryset(self):
        # Filtramos solo los que tienen el soft-delete activo
        queryset = Activo.all_objects.filter(is_deleted=True).select_related(
            'catalogo__categoria', 'catalogo__marca', 'ubicacion'
        )
        
        # Mantenemos la lógica de búsqueda que ya tienes en lista_activos.html
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(bdo__icontains=q) | 
                Q(numero_serie__icontains=q) | 
                Q(etiqueta__icontains=q)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_lista'] = "Papelera de Activos (Eliminados)"
        return context




class RestaurarActivoView(LoginRequiredMixin, GroupRequiredMixin, View):
    """
    Vista para restaurar un activo eliminado (Soft Delete) y registrar la auditoría.
    """
    group_required = ['ADR', 'Operador ADR']

    def post(self, request, pk):
        # 1. Obtener el activo usando el manager que incluye eliminados
        activo = get_object_or_404(Activo.all_objects, pk=pk)
        
        try:
            with transaction.atomic():
                # 2. Restaurar el estado del activo
                activo.is_deleted = False
                activo.save()
                
                # 3. Generar el registro de auditoría manual para la RESTAURACIÓN
                AuditoriaActivo.objects.create(
                    usuario=request.user,
                    accion=AuditoriaActivo.TipoAccion.RESTAURACION,
                    content_type=ContentType.objects.get_for_model(Activo),
                    object_id=activo.id,
                    valor_nuevo=f"Activo restaurado: {activo}"
                )
            
            messages.success(
                request, 
                f"El equipo {activo.catalogo} ({activo.numero_serie or activo.etiqueta}) ha sido restaurado."
            )
        except Exception as e:
            messages.error(request, f"Error técnico al restaurar: {str(e)}")

        return redirect('lista_eliminados')




class ListaCatalogoView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = Catalogo
    template_name = 'inventario/pages/lista_catalogos.html'
    context_object_name = 'catalogos'
    paginate_by = 20
    group_required = ['ADR', 'Alumno en Práctica', 'Auxiliar Operador ADR', 'Operador ADR']

    def get_queryset(self):
        # 1. Anotamos el conteo de activos vinculados que NO están eliminados (is_deleted=False)
        queryset = Catalogo.objects.select_related('categoria', 'marca').annotate(
            activos_count=Count('activo', filter=Q(activo__is_deleted=False))
        )

        # 2. Captura de parámetros
        self.search_query = self.request.GET.get('search', '').strip()
        self.categoria_id = self.request.GET.get('categoria', '')
        self.marca_id = self.request.GET.get('marca', '')
        self.orden = self.request.GET.get('orden', 'nombre')
        self.show_empty = self.request.GET.get('show_empty') == 'on'

        # 3. Filtro por defecto: Solo catálogos con activos (a menos que se marque el checkbox)
        if not self.show_empty:
            queryset = queryset.filter(activos_count__gt=0)

        # 4. Aplicación de otros filtros
        if self.categoria_id:
            queryset = queryset.filter(categoria_id=self.categoria_id)
        if self.marca_id:
            queryset = queryset.filter(marca_id=self.marca_id)
        if self.search_query:
            queryset = queryset.filter(modelo__icontains=self.search_query)

        # 5. Lógica de Ordenamiento
        # 'nombre' ordena por la jerarquía lógica de la composición del nombre
        if self.orden == 'nombre':
            queryset = queryset.order_by('categoria__nombre', 'marca__nombre', 'modelo')
        elif self.orden == '-created_at':
            queryset = queryset.order_by('-created_at')
        elif self.orden == '-updated_at':
            queryset = queryset.order_by('-updated_at')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Preservar filtros en la paginación
        get_copy = self.request.GET.copy()
        if 'page' in get_copy:
            get_copy.pop('page')
        
        context.update({
            'titulo_lista': 'Catálogo de Productos',
            'categorias': Categoria.objects.all().order_by('nombre'),
            'marcas': Marca.objects.all().order_by('nombre'),
            'search_query': self.search_query,
            'categoria_seleccionada': self.categoria_id,
            'marca_seleccionada': self.marca_id,
            'orden_seleccionado': self.orden,
            'show_empty': self.show_empty,
            'query_string': f"&{get_copy.urlencode()}" if get_copy else ""
        })
        return context




class EditarCatalogoView(LoginRequiredMixin, GroupRequiredMixin, UpdateView):
    model = Catalogo
    form_class = CatalogoForm
    template_name = 'inventario/pages/editar_catalogo.html'
    context_object_name = 'catalogo'
    group_required = ['ADR', 'Auxiliar Operador ADR', 'Operador ADR']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_form'] = f"Editar Producto: {self.object.marca.nombre} {self.object.modelo}"
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'El catálogo "{self.object}" ha sido actualizado correctamente.')
        return response

    def form_invalid(self, form):
        messages.error(self.request, 'No se pudo guardar. Revisa los errores del formulario.')
        return super().form_invalid(form)

    def get_success_url(self):
        return reverse_lazy('lista_catalogos')




class CrearCatalogoView(LoginRequiredMixin, GroupRequiredMixin, CreateView):
    model = Catalogo
    form_class = CatalogoForm
    template_name = 'inventario/pages/agregar_catalogo.html'
    group_required = ['ADR', 'Auxiliar Operador ADR', 'Operador ADR']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_form'] = "Registrar Nuevo Producto"
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        # self.object se crea automáticamente en el form_valid de CreateView
        messages.success(self.request, f'El producto "{self.object}" ha sido registrado exitosamente.')
        return response

    def form_invalid(self, form):
        messages.error(self.request, 'No se pudo registrar el producto. Por favor, revisa los errores del formulario.')
        return super().form_invalid(form)

    def get_success_url(self):
        return reverse_lazy('lista_catalogos')




class DetalleCatalogoView(LoginRequiredMixin, GroupRequiredMixin, DetailView):
    model = Catalogo
    template_name = 'inventario/pages/ver_catalogo.html'
    context_object_name = 'catalogo'
    group_required = ['ADR', 'Alumno en Práctica', 'Auxiliar Operador ADR', 'Operador ADR']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Recuperamos los activos vinculados a este catálogo, optimizando las relaciones
        activos_vinculados = self.object.activo_set.select_related(
            'estado', 
            'ubicacion__piso__edificio', 
            'asignado_a'
        ).all().order_by('-updated_at')
        
        context['activos'] = activos_vinculados
        context['total_activos'] = activos_vinculados.count()
        context['titulo_detalle'] = f"Catálogo: {self.object.marca.nombre} {self.object.modelo}"
        
        return context




class EliminarCatalogoView(LoginRequiredMixin, GroupRequiredMixin, DeleteView):
    model = Catalogo
    template_name = 'inventario/pages/eliminar_catalogo.html'
    success_url = reverse_lazy('lista_catalogos')
    group_required = ['ADR', 'Operador ADR']

    def post(self, request, *args, **kwargs):
        try:
            # Intentamos realizar la eliminación física
            return super().post(request, *args, **kwargs)
        except ProtectedError:
            # Si hay activos vinculados (activos o en papelera), evitamos el borrado
            messages.error(
                request, 
                "No se puede eliminar este producto porque existen equipos físicos asociados a él (activos o en la papelera). "
                "Para mantener la integridad del historial, el catálogo debe permanecer en el sistema."
            )
            return redirect('ver_catalogo', pk=self.get_object().id)

    def form_valid(self, form):
        success_url = self.get_success_url()
        nombre_obj = str(self.get_object())
        self.get_object().delete()
        messages.success(self.request, f'El catálogo "{nombre_obj}" ha sido eliminado correctamente.')
        return redirect(success_url)




class ListaCategoriaView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = Categoria
    template_name = 'inventario/pages/lista_categorias.html'
    context_object_name = 'categorias'
    paginate_by = 12  # Un grid de 3x4 o 4x3 funciona muy bien para imágenes
    group_required = ['ADR', 'Alumno en Práctica', 'Auxiliar Operador ADR', 'Operador ADR']

    def get_queryset(self):
        queryset = Categoria.objects.all().order_by('nombre')
        self.search_query = self.request.GET.get('search', '').strip()
        
        if self.search_query:
            queryset = queryset.filter(nombre__icontains=self.search_query)
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.search_query
        context['titulo_lista'] = "Gestión de Categorías"
        return context


class CrearCategoriaView(LoginRequiredMixin, GroupRequiredMixin, CreateView):
    model = Categoria
    form_class = CategoriaForm
    template_name = 'inventario/pages/agregar_categoria.html'
    group_required = ['ADR', 'Operador ADR']
    success_url = reverse_lazy('lista_categorias')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_form'] = "Registrar Nueva Categoría"
        return context

    def form_valid(self, form):
        messages.success(self.request, f'La categoría "{form.instance.nombre}" ha sido creada exitosamente.')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'No se pudo crear la categoría. Revisa los errores en el formulario.')
        return super().form_invalid(form)


class EditarCategoriaView(LoginRequiredMixin, GroupRequiredMixin, UpdateView):
    model = Categoria
    form_class = CategoriaForm
    template_name = 'inventario/pages/editar_categoria.html'
    group_required = ['ADR', 'Operador ADR']
    success_url = reverse_lazy('lista_categorias')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # self.object contiene la instancia de la categoría que se está editando
        context['titulo_form'] = f"Editar Categoría: {self.object.nombre}"
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'La categoría "{self.object.nombre}" se ha actualizado correctamente.')
        return response

    def form_invalid(self, form):
        messages.error(self.request, 'No se pudieron guardar los cambios. Por favor, revisa el formulario.')
        return super().form_invalid(form)


class EliminarCategoriaView(LoginRequiredMixin, GroupRequiredMixin, DeleteView):
    model = Categoria
    template_name = 'inventario/pages/eliminar_categoria.html'
    success_url = reverse_lazy('lista_categorias')
    group_required = ['ADR', 'Operador ADR']

    def post(self, request, *args, **kwargs):
        try:
            # Intentamos la eliminación física
            return super().post(request, *args, **kwargs)
        except ProtectedError:
            # Si existen catálogos asociados, mostramos el error
            messages.error(
                request, 
                f"No se puede eliminar la categoría '{self.get_object().nombre}' porque tiene productos "
                f"registrados en el catálogo. Primero debes eliminar o reasignar dichos productos."
            )
            return redirect('lista_categorias')

    def form_valid(self, form):
        nombre_obj = self.get_object().nombre
        response = super().form_valid(form)
        messages.success(self.request, f'La categoría "{nombre_obj}" ha sido eliminada correctamente.')
        return response


class DetalleCategoriaView(LoginRequiredMixin, GroupRequiredMixin, DetailView):
    model = Categoria
    template_name = 'inventario/pages/ver_categoria.html'
    context_object_name = 'categoria'
    group_required = ['ADR', 'Alumno en Práctica', 'Auxiliar Operador ADR', 'Operador ADR']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        categoria = self.object

        # 1. Queryset base de activos vinculados a esta categoría
        # El manager por defecto de Activo ya filtra is_deleted=False
        activos_qs = Activo.objects.filter(catalogo__categoria=categoria)

        # 2. Cifra total de activos
        total_activos = activos_qs.count()

        # 3. Métricas: Distribución por Estado
        # Esto genera una lista de diccionarios: [{'estado__nombre': 'OPERATIVO', 'total': 15}, ...]
        metricas_estado = activos_qs.values('estado__nombre').annotate(
            total=Count('id')
        ).order_by('-total')

        # 4. Métricas: Distribución por Edificio (Dónde se concentran)
        metricas_edificio = activos_qs.values('ubicacion__piso__edificio__nombre').annotate(
            total=Count('id')
        ).order_by('-total')

        # 5. Listado de Catálogos (Modelos) específicos de esta categoría
        # Aprovechamos de contar cuántos activos tiene cada modelo
        catalogos_relacionados = Catalogo.objects.filter(categoria=categoria).annotate(
            total_equipos=Count('activo', filter=Q(activo__is_deleted=False))
        ).order_by('marca__nombre', 'modelo')

        context.update({
            'titulo_detalle': f"Resumen de Categoría: {categoria.nombre}",
            'total_activos': total_activos,
            'metricas_estado': metricas_estado,
            'metricas_edificio': metricas_edificio,
            'catalogos': catalogos_relacionados,
        })
        return context




class ListaAreaAdministrativaView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = AreaAdministrativa
    template_name = 'inventario/pages/lista_areas.html'
    context_object_name = 'areas'
    paginate_by = 20
    group_required = ['ADR', 'Auxiliar Operador ADR', 'Operador ADR']

    def get_queryset(self):
        queryset = AreaAdministrativa.objects.annotate(
            funcionarios_count=Count('funcionario')
        ).order_by('nombre')
        
        search_query = self.request.GET.get('search', '')
        if search_query:
            # Busca coincidencias en nombre O en siglas
            queryset = queryset.filter(
                Q(nombre__icontains=search_query) | Q(siglas__icontains=search_query)
            )
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        search_query = self.request.GET.get('search', '')
        context['titulo_lista'] = "Áreas Administrativas"
        context['search_query'] = search_query
        context['query_string'] = f"&search={search_query}" if search_query else ""
        return context


class CrearAreaAdministrativaView(LoginRequiredMixin, GroupRequiredMixin, CreateView):
    model = AreaAdministrativa
    form_class = AreaAdministrativaForm
    template_name = 'inventario/pages/agregar_area.html'
    success_url = reverse_lazy('lista_areas')
    group_required = ['ADR', 'Operador ADR']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Nueva Área Administrativa"
        return context

    def form_valid(self, form):
        # Normalizamos la sigla a mayúsculas antes de guardar
        if form.instance.sigla:
            form.instance.sigla = form.instance.sigla.strip().upper()
        
        messages.success(self.request, f"Área '{form.instance.nombre}' creada correctamente.")
        return super().form_valid(form)


class EditarAreaAdministrativaView(LoginRequiredMixin, GroupRequiredMixin, UpdateView):
    model = AreaAdministrativa
    form_class = AreaAdministrativaForm
    template_name = 'inventario/pages/editar_area.html'
    success_url = reverse_lazy('lista_areas')
    group_required = ['ADR', 'Operador ADR']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Editar Área Administrativa"
        # Pasamos el objeto actual para mostrar el nombre antiguo en algún subtítulo
        context['area'] = self.object 
        return context

    def form_valid(self, form):
        # Mantenemos la normalización de la sigla en mayúsculas al editar
        if form.instance.sigla:
            form.instance.sigla = form.instance.sigla.strip().upper()
        
        messages.success(self.request, f"Área '{form.instance.nombre}' actualizada correctamente.")
        return super().form_valid(form)


class EliminarAreaAdministrativaView(LoginRequiredMixin, GroupRequiredMixin, DeleteView):
    model = AreaAdministrativa
    template_name = 'inventario/pages/eliminar_area.html'
    success_url = reverse_lazy('lista_areas')
    context_object_name = 'area'
    group_required = ['ADR', 'Operador ADR']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Eliminar Área Administrativa"
        
        # Enviamos el conteo a la plantilla para adaptar la interfaz
        context['cant_funcionarios'] = self.object.funcionario_set.count()
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        cant_funcionarios = self.object.funcionario_set.count()
        
        # Alerta de seguridad por si intentan saltarse la interfaz
        if cant_funcionarios > 0:
            messages.error(
                request, 
                f"No se puede eliminar '{self.object.nombre}' porque tiene {cant_funcionarios} funcionario(s) asociado(s)."
            )
            return redirect('lista_areas')
        
        # Si no tiene dependencias, procedemos con el flujo normal
        nombre_area = self.object.nombre
        response = super().post(request, *args, **kwargs)
        messages.success(request, f"Área '{nombre_area}' eliminada con éxito.")
        return response
    

class DetalleAreaAdministrativaView(LoginRequiredMixin, GroupRequiredMixin, DetailView):
    model = AreaAdministrativa
    template_name = 'inventario/pages/ver_area.html'
    context_object_name = 'area'
    group_required = ['ADR', 'Auxiliar Operador ADR', 'Operador ADR']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_vista'] = "Detalle de Área Administrativa"
        
        # Recuperamos los funcionarios de esta área ordenados alfabéticamente
        funcionarios = self.object.funcionario_set.select_related('cargo').order_by('nombre')
        
        context['funcionarios_asociados'] = funcionarios
        context['cant_funcionarios'] = funcionarios.count()
        return context




class ListaCargoView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = Cargo
    template_name = 'inventario/pages/lista_cargos.html'
    context_object_name = 'cargos'
    paginate_by = 20
    group_required = ['ADR', 'Auxiliar Operador ADR', 'Operador ADR']

    def get_queryset(self):
        # Cargos ordenados alfabéticamente por nombre de forma ascendente
        queryset = Cargo.objects.annotate(
            funcionarios_count=Count('funcionario')
        ).order_by('nombre')
        
        # Filtro de búsqueda por nombre de cargo
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(nombre__icontains=search_query)
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        search_query = self.request.GET.get('search', '')
        
        context['titulo_lista'] = "Cargos de Funcionarios"
        context['search_query'] = search_query
        context['query_string'] = f"&search={search_query}" if search_query else ""
        return context


class CrearCargoView(LoginRequiredMixin, GroupRequiredMixin, CreateView):
    model = Cargo
    form_class = CargoForm
    template_name = 'inventario/pages/agregar_cargo.html'
    success_url = reverse_lazy('lista_cargos')
    group_required = ['ADR', 'Operador ADR']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Nuevo Cargo"
        return context

    def form_valid(self, form):
        messages.success(self.request, f"Cargo '{form.instance.nombre}' creado correctamente.")
        return super().form_valid(form)


class EditarCargoView(LoginRequiredMixin, GroupRequiredMixin, UpdateView):
    model = Cargo
    form_class = CargoForm
    template_name = 'inventario/pages/editar_cargo.html'
    success_url = reverse_lazy('lista_cargos')
    context_object_name = 'cargo'
    group_required = ['ADR', 'Operador ADR']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Editar Cargo"
        return context

    def form_valid(self, form):
        messages.success(self.request, f"Cargo '{form.instance.nombre}' actualizado correctamente.")
        return super().form_valid(form)


class EliminarCargoView(LoginRequiredMixin, GroupRequiredMixin, DeleteView):
    model = Cargo
    template_name = 'inventario/pages/eliminar_cargo.html'
    success_url = reverse_lazy('lista_cargos')
    context_object_name = 'cargo'
    group_required = ['ADR', 'Operador ADR']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Eliminar Cargo"
        # Contamos cuántos funcionarios tienen este cargo asignado
        context['cant_funcionarios'] = self.object.funcionario_set.count()
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        cant_funcionarios = self.object.funcionario_set.count()
        
        # Si un usuario malintencionado intenta forzar el POST por consola teniendo dependencias:
        if cant_funcionarios > 0:
            messages.error(
                request, 
                f"No se puede eliminar el cargo '{self.object.nombre}' porque tiene {cant_funcionarios} funcionario(s) asociado(s)."
            )
            return redirect('lista_cargos')
        
        # Si está libre de funcionarios, procedemos con la eliminación estándar
        nombre_cargo = self.object.nombre
        response = super().post(request, *args, **kwargs)
        messages.success(request, f"Cargo '{nombre_cargo}' eliminado con éxito.")
        return response




class ListaFuncionarioView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = Funcionario
    template_name = 'inventario/pages/lista_funcionarios.html'
    context_object_name = 'funcionarios'
    paginate_by = 10
    group_required = ['ADR', 'Auxiliar Operador ADR', 'Operador ADR']

    def get_queryset(self):
        # Ordenamos los funcionarios alfabéticamente por nombre
        queryset = Funcionario.objects.select_related('cargo', 'area').annotate(
            activos_count=Count('activo')
        ).order_by('nombre')
        
        # Captura de parámetros GET
        search_query = self.request.GET.get('search', '')
        area_filtrada = self.request.GET.get('area', '')
        cargo_filtrado = self.request.GET.get('cargo', '')

        # 1. Aplicar Búsqueda por Texto (Filtra por nombre del funcionario)
        if search_query:
            queryset = queryset.filter(nombre__icontains=search_query)

        # 2. Aplicar Filtro Avanzado por Área Administrativa
        if area_filtrada:
            queryset = queryset.filter(area_id=area_filtrada)

        # 3. Aplicar Filtro Avanzado por Cargo
        if cargo_filtrado:
            queryset = queryset.filter(cargo_id=cargo_filtrado)
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Datos estáticos para los dropdowns de los filtros
        context['areas'] = AreaAdministrativa.objects.all().order_by('nombre')
        context['cargos'] = Cargo.objects.all().order_by('nombre')
        
        # Mantener los estados de los filtros en la plantilla
        context['titulo_lista'] = "Nómina de Funcionarios"
        context['search_query'] = self.request.GET.get('search', '')
        context['area_seleccionada'] = self.request.GET.get('area', '')
        context['cargo_seleccionado'] = self.request.GET.get('cargo', '')
        
        # Reconstruir la query string para que la paginación no rompa los filtros
        params = []
        if context['search_query']: params.append(f"search={context['search_query']}")
        if context['area_seleccionada']: params.append(f"area={context['area_seleccionada']}")
        if context['cargo_seleccionado']: params.append(f"cargo={context['cargo_seleccionado']}")
        
        context['query_string'] = f"&{'&'.join(params)}" if params else ""
        return context


class CrearFuncionarioView(LoginRequiredMixin, GroupRequiredMixin, CreateView):
    model = Funcionario
    form_class = FuncionarioForm
    template_name = 'inventario/pages/agregar_funcionario.html'
    success_url = reverse_lazy('lista_funcionarios')
    group_required = ['ADR', 'Operador ADR']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Nuevo Funcionario"
        return context

    def form_valid(self, form):
        # Guardamos el objeto primero para que ejecute el clean() del modelo
        response = super().form_valid(form)
        # Notificamos con el nombre ya normalizado en mayúsculas
        messages.success(self.request, f"Funcionario '{self.object.nombre}' registrado correctamente.")
        return response




class EditarFuncionarioView(LoginRequiredMixin, GroupRequiredMixin, UpdateView):
    model = Funcionario
    form_class = FuncionarioForm
    template_name = 'inventario/pages/editar_funcionario.html'
    success_url = reverse_lazy('lista_funcionarios')
    context_object_name = 'funcionario'
    group_required = ['ADR', 'Operador ADR']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Editar Funcionario"
        return context

    def form_valid(self, form):
        # Al guardar, el método clean() del modelo se ejecuta automáticamente
        response = super().form_valid(form)
        messages.success(self.request, f"Datos de '{self.object.nombre}' actualizados correctamente.")
        return response


class EliminarFuncionarioView(LoginRequiredMixin, GroupRequiredMixin, DeleteView):
    model = Funcionario
    template_name = 'inventario/pages/eliminar_funcionario.html'
    success_url = reverse_lazy('lista_funcionarios')
    context_object_name = 'funcionario'
    group_required = ['ADR', 'Operador ADR']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_formulario'] = "Eliminar Funcionario"
        # Contamos cuántos activos vigentes tiene asignados este funcionario
        context['cant_activos'] = self.object.activo_set.count()
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        cant_activos = self.object.activo_set.count()
        
        # Guardaparques de backend en caso de solicitudes post maliciosas o concurrentes
        if cant_activos > 0:
            messages.error(
                request, 
                f"Operación denegada: '{self.object.nombre}' mantiene {cant_activos} activo(s) a su cargo."
            )
            return redirect('lista_funcionarios')
        
        nombre_funcionario = self.object.nombre
        response = super().post(request, *args, **kwargs)
        messages.success(request, f"Funcionario '{nombre_funcionario}' eliminado correctamente de la nómina.")
        return response


class DetalleFuncionarioView(LoginRequiredMixin, GroupRequiredMixin, DetailView):
    model = Funcionario
    template_name = 'inventario/pages/ver_funcionario.html'
    context_object_name = 'funcionario'
    group_required = ['ADR', 'Auxiliar Operador ADR', 'Operador ADR']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo_vista'] = "Ficha del Funcionario"
        
        # Optimizamos agregando 'catalogo__categoria' y 'catalogo__marca'
        activos = self.object.activo_set.select_related(
            'catalogo__categoria', 
            'catalogo__marca', 
            'estado', 
            'ubicacion'
        ).order_by('-created_at')
        
        context['activos_asignados'] = activos
        context['cant_activos'] = activos.count()
        return context